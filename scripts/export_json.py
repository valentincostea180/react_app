import openpyxl
import os
import logging
from datetime import datetime, timedelta
import sqlite3
import json
from openpyxl.utils import column_index_from_string


# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


# Constants
EXCEL_FILE = "./public/uploads/excelMare.xlsx"
PRODUCTION_FILE = "./public/uploads/productionPlan.xlsx"
BOS_FILE = "./public/uploads/Formular BOS.xlsx"
NM_FILE = "./public/uploads/Formular raportare eveniment la limita producerii unui accident sau situație periculoasă.xslx"
DB_FILE = "./public/data/production_data.db"


WEEKLY_OUTPUT = "./public/data/weekly_summary.json"
NM_OUTPUT = "./public/data/NM.json"
BOS_OUTPUT = "./public/data/BOS.json"
TARGET_OUTPUT = "./public/data/target.json"
NEWS_OUTPUT = "./public/data/news.json"


if not os.path.exists(EXCEL_FILE):
    raise FileNotFoundError(f"Fișierul Excel nu a fost găsit: {EXCEL_FILE}")


if not os.path.exists(PRODUCTION_FILE):
    raise FileNotFoundError(f"Fișierul Production Plan nu a fost găsit: {PRODUCTION_FILE}")


if not os.path.exists(BOS_FILE):
    raise FileNotFoundError(f"Fișierul BOS nu a fost găsit: {BOS_FILE}")


def last_day_month(actual_day):
    """Calculate the last day of the month for a given date."""
    if isinstance(actual_day, str):
        actual_day = datetime.strptime(actual_day, "%Y%m%d")
   
    next_month = actual_day.replace(day=28) + timedelta(days=4)
    return (next_month - timedelta(days=next_month.day)).date()


# --- Generic data extraction function for operational and minor stoppages ---
def extract_data(sheet, header_row, first_data_row, col_start, col_end, comment_col):
    """Return a list of rows grouped by their date."""
    data_by_date = {}


    for row in range(first_data_row, sheet.max_row + 1):
        date_value = sheet.cell(row=row, column=2).value  # Column B
        if not isinstance(date_value, datetime):
            continue
        date_key = date_value.strftime("%Y%m%d")


        if date_key not in data_by_date:
            data_by_date[date_key] = {"rows": [], "total_min": 0}


        for col in range(col_start, col_end + 1):
            value = sheet.cell(row=row, column=col).value
            if isinstance(value, (int, float)) and value != 0:
                equipment_name = sheet.cell(row=header_row, column=col).value
                comment = sheet.cell(row=row, column=comment_col).value
                row_data = {
                    "schimb": sheet.cell(row=row, column=3).value,
                    "produs": sheet.cell(row=row, column=4).value,
                    "nume_echipament": equipment_name,
                    "minute": value,
                    "comentariu": str(comment) if comment else ""
                }
                data_by_date[date_key]["rows"].append(row_data)
                data_by_date[date_key]["total_min"] += value


    return data_by_date


# extragere data windwos log viewer
def extract_section_manager_production():
    """
    Extract production data from the Data sheet - simple version
    """
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet_name = "Data"
       
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Lipsă foaie: {sheet_name}")
       
        sheet = wb[sheet_name]
        data_by_date = {}
       
        for row in range(11, sheet.max_row + 1):
            date_value = sheet.cell(row=row, column=2).value  # Column B (Date)
           
            # Skip rows without valid date
            if not isinstance(date_value, datetime):
                continue
           
            date_key = date_value.strftime("%Y%m%d")
           
            if date_key not in data_by_date:
                data_by_date[date_key] = []
           
            # Extract all columns for this row
            row_data = {}
           
            for col in range(1, 53):  # Columns to AZ
                value = sheet.cell(row=row, column=col).value
                header_name = sheet.cell(row=9, column=col).value
               
                # Handle error values
                if value is not None and str(value).startswith('#'):
                    value = None
               
                row_data[header_name] = value
           
            data_by_date[date_key].append(row_data)
       
        logging.info(f"Successfully extracted production data for {len(data_by_date)} dates")
        return data_by_date
       
    except Exception as e:
        logging.error(f"Error extracting production data: {e}")
        raise


# --- Function for GE data ---
def extract_inceput_data(sheet):
    data_by_date = {}
    target_row = 11
    for row in range(target_row, sheet.max_row + 1):
        date_value = sheet.cell(row=row, column=2).value
        if not isinstance(date_value, datetime):
            continue
        date_key = date_value.strftime("%Y%m%d")
        if date_key not in data_by_date:
            data_by_date[date_key] = []


        for col in range(3, 30):  # Columns C to AC
            value = sheet.cell(row=row, column=col).value
            operation_name = sheet.cell(row=8, column=col).value
            if value is not None:
                if isinstance(value, datetime):
                    value = value.strftime("%H:%M:%S")
                data_by_date[date_key].append({
                    "operatie": operation_name,
                    "valoare": value
                })
    return data_by_date


# --- Function to extract production plan data ---
def extract_production_plan_data():
    """Extract production plan data from Excel file."""
    try:
        wb = openpyxl.load_workbook(PRODUCTION_FILE, data_only=True)


        sheet_name = "Line 3"


        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Lipsă foaie: {sheet_name}")
       
        sheet = wb[sheet_name]
       
        data_by_date = {}
       
        # Define the columns we need
        columns_needed = {
            'D': 'cod_sap',           # Column D
            'E': 'tip_produs',        # Column E
            'F': 'aroma',             # Column F
            'G': 'tray',              # Column G
            'H': 'gramaj',            # Column H
            'K': 'pcs_bax',           # Column K
            'N': 'comanda_initiala',  # Column N
            'Q': 'start_date',        # Column Q
            'R': 'end_date',          # Column R
            'S': 'ore_productie',     # Column S
            'T': 'shifturi'           # Column T
        }
       
        # Create a mapping of column letters to our needed fields
        col_mapping = {}
        for col_letter, field_name in columns_needed.items():
            col_index = openpyxl.utils.column_index_from_string(col_letter)
            col_mapping[col_index] = field_name
       
        # Extract data from each row
        for row in range(2, sheet.max_row + 1):  # Start from row 2 (assuming headers in row 1)
            # Check if we have a date in column Q (Start Date)
            date_value = sheet.cell(row=row, column=openpyxl.utils.column_index_from_string('Q')).value
            if not isinstance(date_value, datetime):
                continue
               
            date_key = date_value.strftime("%Y%m%d")
           
            if date_key not in data_by_date:
                data_by_date[date_key] = []
           
            # Extract the needed columns
            row_data = {}
            for col_index, field_name in col_mapping.items():
                value = sheet.cell(row=row, column=col_index).value
                row_data[field_name] = value
           
            data_by_date[date_key].append(row_data)
       
        return data_by_date
    except Exception as e:
        logging.error(f"Eroare la extragerea datelor din production plan: {e}")
        raise


# --- Function to extract BOS data ---
def extract_bos_data():
    """Extract BOS data from Excel file."""
    try:
        wb = openpyxl.load_workbook(BOS_FILE, data_only=True)
        sheet_name = "Report total safe-unsafe"


        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Lipsă foaie: {sheet_name}")
       
        sheet = wb[sheet_name]
        current_date = datetime.now().strftime("%Y%m%d")
       
        # Extract the values from specific cells
        actiuni_nesigure = sheet["B17"].value or 0
        actiuni_sigure = sheet["C17"].value or 0
       
        return {
            "actiuni_sigure": int(actiuni_sigure) if actiuni_sigure else 0,
            "actiuni_nesigure": int(actiuni_nesigure) if actiuni_nesigure else 0
        }
       
    except Exception as e:
        logging.error(f"Eroare la extragerea datelor din BOS: {e}")
        raise


def extract_nm_data():
    """Extract NM data from Excel file."""
    try:
        wb = openpyxl.load_workbook(NM_FILE, data_only=True)
        sheet_name = "Total reports"


        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Lipsă foaie: {sheet_name}")
       
        sheet = wb[sheet_name]
        current_date = datetime.now().strftime("%Y%m%d")


        sheet["B2"] = current_date
       
        # Extract the values from specific cells
        nm = sheet["B6"].value or 0
        uc = sheet["B7"].value or 0
       
        return {
            "near_miss": int(nm) if nm else 0,
            "unsafe_condition": int(uc) if uc else 0
        }
       
    except Exception as e:
        logging.error(f"Eroare la extragerea datelor din NM: {e}")
        raise


# --- Function to extract target data ---
def extract_target_data():
    """Extract target data from Excel file."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet_name = "Monthly"


        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Lipsă foaie: {sheet_name}")
       
        sheet = wb[sheet_name]
        now = datetime.now()
        last_day = last_day_month(now)
       
        target_ge = 80  # Default value
       
        for row in range(4, sheet.max_row + 1):
            date_value = sheet.cell(row=row, column=2).value
            if isinstance(date_value, datetime) and date_value.date() == last_day:
                target_ge = sheet.cell(row=row, column=76).value or 80 # Column BX
                break
       
        return [{
            "operation": "GE",
            "target": int(target_ge*100) if target_ge else 80
        },
        {
            "operation": "Waste",
            "target": 3
        },
        {
            "operation": "Volume Produced",
            "target": 20
        },
        {
            "operation": "Speed Loss",
            "target": 1
        }]
       
    except Exception as e:
        logging.error(f"Eroare la extragerea datelor din target: {e}")
        raise


def extract_news_data():
    """Extract target data from Excel file."""
    try:
        return [{
            "message": "NU VA UITATI ANITIFOANELE",
        },
        {
            "message": "TRAINING OPL, ORA 13:00",
        },
        {
            "message": "VERIFICATI STARE MASINI INAINTE DE TURA",
        },
        {
            "message": "RAPORTATI ORICE PROBLEMA IMEDIAT",
        },
        {
            "message": "MENTINETI ZONA DE LUCRU CURATA",
        }
        ]
       
    except Exception as e:
        logging.error(f"Eroare la extragere: {e}")
        raise


def extract_weekly_data():
    """Extract weekly summary data from Excel file."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet_name = "Weekly"
       
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Lipsă foaie: {sheet_name}")
       
        sheet = wb[sheet_name]
        weekly_data = []
       
        # Assuming the data starts from row 4 (with headers in row 3)
        for row in range(4, sheet.max_row + 1):
            week_number = sheet.cell(row=row, column=1).value  # Column A
            ge_value = sheet.cell(row=row, column=26).value    # Column Z
            produced_volume = sheet.cell(row=row, column=27).value  # Column AA
            breakdowns = sheet.cell(row=row, column=17).value  # Column Q
            operational_losses = sheet.cell(row=row, column=18).value  # Column R
            waste = sheet.cell(row=row, column=32).value       # Column AF
            speed_loss = sheet.cell(row=row, column=23).value  # Column G
           
            # Only add rows with valid week numbers
            if week_number and isinstance(week_number, (int, float)):
                weekly_data.append({
                    "week_number": int(week_number),
                    "ge": round(float(ge_value), 2) if ge_value else 0,
                    "produced_volume": round(float(produced_volume), 2) if produced_volume else 0,
                    "breakdowns": round(float(breakdowns), 2 ) if breakdowns else 0,
                    "operational_losses": round(float(operational_losses), 2) if operational_losses else 0,
                    "waste": round(float(waste), 2) if waste else 0,
                    "speed_loss": round(float(speed_loss), 2) if speed_loss else 0
                })
       
        return weekly_data
       
    except Exception as e:
        logging.error(f"Eroare la extragerea datelor săptămânale: {e}")
        raise


# --- Extract all data from both Excel files ---
def extract_data_from_all_files():
    try:
        # Process excelMare.xlsx
        wb_main = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        required_sheets = ["Operational Losses", "Minor stoppages", "Daily_CALC", "Breakdowns"]
        for sheet_name in required_sheets:
            if sheet_name not in wb_main.sheetnames:
                raise ValueError(f"Lipsă foaie în excelMare: {sheet_name}")


        opp_data = extract_data(wb_main["Operational Losses"], 9, 11, 10, 36, 44)
        minor_data = extract_data(wb_main["Minor stoppages"], 9, 11, 10, 40, 47)
        break_data = extract_data(wb_main["Breakdowns"], 9, 11, 10, 40, 47)
        inceput_data = extract_inceput_data(wb_main["Daily_CALC"])
       
        # Process productionPlan.xlsx
        production_plan_data = extract_production_plan_data()


        section_manager_data = extract_section_manager_production()


        # Combine all dates into a single dict
        all_dates = {}
        all_keys = set(list(opp_data.keys()) + list(minor_data.keys()) +
                      list(inceput_data.keys()) + list(production_plan_data.keys()) )
       
        for date_key in all_keys:
            all_dates[date_key] = {
                "metadata": {
                    "total_opp": opp_data.get(date_key, {}).get("total_min", 0),
                    "total_minor": minor_data.get(date_key, {}).get("total_min", 0)
                },
                "operational_losses": opp_data.get(date_key, {}).get("rows", []),
                "minor_stoppages": minor_data.get(date_key, {}).get("rows", []),
                "breakdowns": break_data.get(date_key, {}).get("rows", []),
                "inceput": inceput_data.get(date_key, []),
                "production_plan": production_plan_data.get(date_key, []),
                "section_manager_production": section_manager_data.get(date_key, []),
                "rapoarte": []
            }
        return all_dates
    except Exception as e:
        logging.error(f"Eroare la extragerea datelor: {e}")
        raise


# --- JSON functions ---
def save_to_json(data, filename):
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        logging.info(f"Datele au fost salvate în {filename}")
    except Exception as e:
        logging.error(f"Eroare la salvarea datelor în JSON: {e}")
        raise


def is_file_empty_or_blank(filename):
    if not os.path.exists(filename):
        return True
    return os.path.getsize(filename) == 0


# --- SQLite3 functions ---
def create_tables(conn):
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS metadata (
            date TEXT PRIMARY KEY,
            total_opp INTEGER,
            total_minor INTEGER
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS operational_losses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            schimb INTEGER,
            produs TEXT,
            nume_echipament TEXT,
            minute INTEGER,
            comentariu TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS minor_stoppages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            schimb INTEGER,
            produs TEXT,
            nume_echipament TEXT,
            minute INTEGER,
            comentariu TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS section_manager_production (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            week TEXT,
            shift TEXT,
            maximum_time_hours REAL,
            official_holidays REAL,
            no_demand REAL,
            force_majeure REAL,
            hours_used REAL,
            product_description TEXT,
            maximum_theoretical_capacity REAL,
            team_leader_name TEXT,
            volume_produced REAL,
            waste REAL,
            rework REAL,
            rework_used_same_shift REAL,
            gross_waste REAL,
            standard_weight REAL,
            average_weight REAL,
            overweight REAL,
            head_count_actual INTEGER,
            head_count_standard INTEGER,
            changeover_time_standard REAL,
            production_hours REAL,
            planned_maintenance REAL,
            planned_autonomous_maintenance REAL,
            sanitation REAL,
            changeovers REAL,
            planned_stops REAL,
            consumables_replacement REAL,
            start_finish_production REAL,
            minor_stoppages REAL,
            breakdowns REAL,
            operational_losses REAL,
            line_delays REAL,
            labor_management_losses REAL,
            material_shortages REAL,
            quality_loss REAL,
            speed_loss REAL,
            defect_free_full_speed_time REAL,
            oee REAL,
            ge REAL,
            dvl_consumption REAL,
            productivity REAL,
            overweight_2 REAL,
            yield_loss REAL,
            yield_loss_2 REAL,
            sortiment_modifications REAL,
            cumulative_changeover_ytd REAL,
            batch TEXT,
            best_before_day TEXT,
            sap_code TEXT,
            po TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS breakdowns (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            schimb INTEGER,
            produs TEXT,
            nume_echipament TEXT,
            minute INTEGER,
            comentariu TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS inceput (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            operatie TEXT,
            valoare TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS rapoarte (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            oraIn TEXT,
            oraOut TEXT,
            zona TEXT,
            tip TEXT,
            tipSecundar TEXT,
            masina TEXT,
            ansamblu TEXT,
            problema TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS production_plan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT,
            cod_sap TEXT,
            tip_produs TEXT,
            aroma TEXT,
            tray TEXT,
            gramaj REAL,
            pcs_bax INTEGER,
            comanda_initiala REAL,
            start_date TEXT,
            end_date TEXT,
            ore_productie REAL,
            shifturi INTEGER
        )
    """)
    conn.commit()


def insert_data(conn, data):
    cur = conn.cursor()
    for date_key, content in data.items():
        # --- clear old rows for this date (but NOT rapoarte) ---
        cur.execute("DELETE FROM metadata WHERE date = ?", (date_key,))
        cur.execute("DELETE FROM operational_losses WHERE date = ?", (date_key,))
        cur.execute("DELETE FROM minor_stoppages WHERE date = ?", (date_key,))
        cur.execute("DELETE FROM breakdowns WHERE date = ?", (date_key,))
        cur.execute("DELETE FROM inceput WHERE date = ?", (date_key,))
        cur.execute("DELETE FROM production_plan WHERE date = ?", (date_key,))
        cur.execute("DELETE FROM section_manager_production WHERE date = ?", (date_key,))


        # metadata
        cur.execute("""
            INSERT INTO metadata (date, total_opp, total_minor)
            VALUES (?, ?, ?)
        """, (date_key, content["metadata"]["total_opp"], content["metadata"]["total_minor"]))


        # operational losses
        for row in content["operational_losses"]:
            cur.execute("""
                INSERT INTO operational_losses (date, schimb, produs, nume_echipament, minute, comentariu)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (date_key, row["schimb"], row["produs"], row["nume_echipament"], row["minute"], row["comentariu"]))


        # minor stoppages
        for row in content["minor_stoppages"]:
            cur.execute("""
                INSERT INTO minor_stoppages (date, schimb, produs, nume_echipament, minute, comentariu)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (date_key, row["schimb"], row["produs"], row["nume_echipament"], row["minute"], row["comentariu"]))


        # breakdowns
        for row in content["breakdowns"]:
            cur.execute("""
                INSERT INTO breakdowns (date, schimb, produs, nume_echipament, minute, comentariu)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (date_key, row["schimb"], row["produs"], row["nume_echipament"], row["minute"], row["comentariu"]))


        # inceput
        for row in content["inceput"]:
            cur.execute("""
                INSERT INTO inceput (date, operatie, valoare)
                VALUES (?, ?, ?)
            """, (date_key, row["operatie"], str(row["valoare"])))


        # production plan
        for row in content["production_plan"]:
            cur.execute("""
                INSERT INTO production_plan
                (date, cod_sap, tip_produs, aroma, tray, gramaj, pcs_bax,
                 comanda_initiala, start_date, end_date, ore_productie, shifturi)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                date_key,
                row.get("cod_sap"),
                row.get("tip_produs"),
                row.get("aroma"),
                row.get("tray"),
                row.get("gramaj"),
                row.get("pcs_bax"),
                row.get("comanda_initiala"),
                row.get("start_date"),
                row.get("end_date"),
                row.get("ore_productie"),
                row.get("shifturi")
            ))


        for row in content.get("section_manager_production", []):
            # Helper function to handle numeric values and errors
            def safe_value(value, default=None):
                if value is None or str(value).startswith('#') or value == '#VALUE!':
                    return default
                try:
                    # Handle datetime.time objects
                    if hasattr(value, 'strftime'):  # Check if it's a time/date object
                        return value.strftime('%H:%M:%S')  # Convert to string
                    return float(value) if isinstance(value, (int, float)) else value
                except (ValueError, TypeError):
                    return value


            cur.execute("""
                INSERT INTO section_manager_production (
                    date, week, shift, maximum_time_hours, official_holidays, no_demand,
                    force_majeure, hours_used, product_description, maximum_theoretical_capacity,
                    team_leader_name, volume_produced, waste, rework, rework_used_same_shift,
                    gross_waste, standard_weight, average_weight, overweight, head_count_actual,
                    head_count_standard, changeover_time_standard, production_hours,
                    planned_maintenance, planned_autonomous_maintenance, sanitation,
                    changeovers, planned_stops, consumables_replacement, start_finish_production,
                    minor_stoppages, breakdowns, operational_losses, line_delays,
                    labor_management_losses, material_shortages, quality_loss, speed_loss,
                    defect_free_full_speed_time, oee, ge, dvl_consumption, productivity,
                    overweight_2, yield_loss, yield_loss_2, sortiment_modifications,
                    cumulative_changeover_ytd, batch, best_before_day, sap_code, po
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                         ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                         ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                date_key,  # date comes first, then the row data
                safe_value(row.get("Week")),
                safe_value(row.get("Shift")),
                safe_value(row.get("Maximum time (Hours)")),
                safe_value(row.get("Official holidays")),
                safe_value(row.get("No demand")),
                safe_value(row.get("Force Majeure")),
                safe_value(row.get("Hours used")),
                safe_value(row.get("Product Description")),
                safe_value(row.get("Maximum theoretical capacity")),
                safe_value(row.get("Team Leader name")),
                safe_value(row.get("Volume Produced")),
                safe_value(row.get("Waste")),
                safe_value(row.get("Rework")),
                safe_value(row.get("Rework used in the same shift")),
                safe_value(row.get("Gross Waste")),
                safe_value(row.get("Standard weight of product")),
                safe_value(row.get("Average weight")),
                safe_value(row.get("Overweight")),
                safe_value(row.get("Head Count - Actual")),
                safe_value(row.get("Head Count - Standard")),
                safe_value(row.get("Changeover time - Standard")),
                safe_value(row.get("Production hours")),
                safe_value(row.get("Planned Maintenance")),
                safe_value(row.get("Planned Autonomous Maintenance")),
                safe_value(row.get("Sanitation")),
                safe_value(row.get("Changeovers")),
                safe_value(row.get("Planned Stops")),
                safe_value(row.get("Consumables replacement")),
                safe_value(row.get("Start and Finish Production")),
                safe_value(row.get("Minor Stoppages")),
                safe_value(row.get("Breakdowns")),
                safe_value(row.get("Operational Losses")),
                safe_value(row.get("Line Delays")),
                safe_value(row.get("Labor Management Losses")),
                safe_value(row.get("Material Shortages")),
                safe_value(row.get("Quality Loss")),
                safe_value(row.get("Speed Loss")),
                safe_value(row.get("Defect-free full speed operating time")),
                safe_value(row.get("OEE")),
                safe_value(row.get("GE")),
                safe_value(row.get("DVL consumption")),
                safe_value(row.get("Productivity")),
                safe_value(row.get("Overweight_2")),  # Use the renamed column
                safe_value(row.get("Yield loss")),
                safe_value(row.get("Yield loss_2")),
                safe_value(row.get("Modificari ale sortimentului de la inceputul anului")),
                safe_value(row.get("Cummulative Changeover Time YTD")),
                safe_value(row.get("Batch")),
                safe_value(row.get("Best Before day")),
                safe_value(row.get("SAP code")),
                safe_value(row.get("PO"))
            ))


        # rapoarte (preserved from previous runs)
        # Note: We don't delete rapoarte to preserve user-added reports
   
    conn.commit()


def to_json(file_path, file_type):
    """Save data to JSON file based on file type."""
    if is_file_empty_or_blank(file_path):
        if file_type == "BOS":
            formatted_data = extract_bos_data()
        elif file_type == "TARGET":
            formatted_data = extract_target_data()
        elif file_type == "NEWS":
            formatted_data = extract_news_data()
        elif file_type == "WEEKLY":
            formatted_data = extract_weekly_data()
        elif file_type == "NM":
            formatted_data = extract_nm_data()
        else:
            logging.error(f"Tip de fișier necunoscut: {file_type}")
            return
           
        logging.info(f"Salvez datele în fișier JSON: {file_path}")
        save_to_json(formatted_data, file_path)
        logging.info("Procesul s-a terminat cu succes!")
        print(f"Fișier generat: {file_path}")
    else:
        print("Fișier existent.")


def to_sqlite3():
    formatted_data = extract_data_from_all_files()
    conn = sqlite3.connect(DB_FILE)
    create_tables(conn)
    insert_data(conn, formatted_data)
    conn.close()
    logging.info("Datele au fost salvate în SQLite!")
    print(f"Datele din fișierele Excel au fost salvate în baza de date.")


def main():
    try:
        logging.info("Încep extragerea datelor din fișierele Excel...")
        to_sqlite3()
        to_json(BOS_OUTPUT, "BOS")
        to_json(TARGET_OUTPUT, "TARGET")
        to_json(NEWS_OUTPUT, "NEWS")
        to_json(WEEKLY_OUTPUT, "WEEKLY")
        to_json(NM_OUTPUT, "NM")
    except Exception as e:
        logging.error(f"A apărut o eroare: {e}")


if __name__ == "__main__":
    main()



